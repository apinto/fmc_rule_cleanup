#!/usr/bin/env python3
"""
FMC Access Rule Hit Count Analysis and Auto-Disable Script

This script analyzes Cisco FMC access control rules based on hit counts
and automatically disables unused rules that meet specific criteria.

Author: Artur Pinto (arturj.pinto@gmail.com)
"""

import argparse
import datetime
import ipaddress
import logging
import sys
import time
from typing import Dict, List, Optional, Set, Union

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not installed. Excel export functionality will be disabled.")

try:
    import fmcapi
    import requests
except ImportError:
    fmcapi = None


# Progress bar function for console output
def print_progress_bar(iteration, total, prefix='', suffix='', length=50, fill='█', print_end="\r"):
    """
    Call in a loop to create a terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str) - can include retry information
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        print_end   - Optional  : end character (e.g. "\r", "\n") (Str)
    """
    # Clear any previous text in the line first (using 120 spaces to ensure full line clearing)
    print("\r" + " " * 120, end="\r")
    
    # Calculate percentage and create bar
    percent = ("{0:.1f}").format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    
    # Format text with counter, percentage and any extra info
    status = f"{iteration}/{total}"
    output = f'\r{prefix} |{bar}| {percent}% {status} {suffix}'
    
    # Print the formatted progress bar
    print(output, end=print_end)
    
    # Print a new line if we're at the end
    if iteration >= total:
        print()


class FMCRuleManager:
    """Manages FMC access rule operations including hit count analysis and rule disabling."""
    
    def __init__(self, host: str, username: str, password: str, device_name: str,
                 autodeploy: bool = False, page_limit: int = 500, debug: bool = False,
                 timeout: int = 10, log_file: Optional[str] = None,
                 max_rules_to_disable: int = 1000, dry_run: bool = False,
                 exclude_zones: Optional[List[str]] = None, year_threshold: int = None,
                 rule_actions: Optional[List[str]] = None, exclude_prefixes: Optional[List[str]] = None,
                 prefix_match_mode: str = 'overlap'):
        """
        Initialize FMC Rule Manager.
        
        Args:
            host: FMC host IP address
            username: FMC username
            password: FMC password  
            device_name: Target device name
            autodeploy: Enable auto-deployment
            page_limit: API page limit for queries
            debug: Enable debug logging
            timeout: API timeout in seconds
            log_file: Log file name (optional)
            max_rules_to_disable: Maximum number of rules to disable per run
            dry_run: If True, simulate actions without making changes
            exclude_zones: List of zones to exclude from rule processing
            year_threshold: Consider rules created before this year for disabling (default: current year - 1)
            rule_actions: List of rule actions to consider (ALLOW, BLOCK, or both)
            exclude_prefixes: List of IP prefixes (CIDR) to exclude from rule processing
            prefix_match_mode: Mode for prefix matching - 'overlap' (any overlap) or 'subnet' (subset only)
        """
        self.host = host
        self.username = username
        self.password = password
        self.device_name = device_name
        self.autodeploy = autodeploy
        self.page_limit = page_limit
        self.debug = debug
        self.timeout = timeout
        self.log_file = log_file
        self.max_rules_to_disable = max_rules_to_disable
        self.dry_run = dry_run
        self.exclude_zones = exclude_zones or []
        # Set default year threshold to previous year if not specified
        self.year_threshold = year_threshold if year_threshold is not None else datetime.datetime.now().year - 1
        self.rule_actions = rule_actions or ['ALLOW']
        self.prefix_match_mode = prefix_match_mode
        
        # Parse and validate exclude prefixes
        self.exclude_prefixes: List[Union[ipaddress.IPv4Network, ipaddress.IPv6Network]] = []
        if exclude_prefixes:
            for prefix in exclude_prefixes:
                try:
                    # Parse as network (handles both single IPs and CIDR notation)
                    network = ipaddress.ip_network(prefix, strict=False)
                    self.exclude_prefixes.append(network)
                except ValueError as e:
                    logging.warning(f"Invalid IP prefix '{prefix}': {e}. Skipping.")
        
        # Cache for resolved network objects to avoid repeated API calls
        self._network_object_cache: Dict[str, List[str]] = {}
        
        # Check if fmcapi is available
        if fmcapi is None:
            raise ImportError("fmcapi module is required. Install with: pip install fmcapi")
        
        # Setup logging
        self._setup_logging()
        
    def _setup_logging(self) -> None:
        """Configure logging to go to file only, not to console."""
        log_level = logging.DEBUG if self.debug else logging.INFO
        
        # Remove any existing handlers to prevent duplicate logging
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        
        if self.log_file:
            # Only log to file, not to console, and append to the log file
            logging.basicConfig(
                level=log_level,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[logging.FileHandler(self.log_file, mode='a')]
            )
            print(f"Logging to file: {self.log_file}")
        else:
            # No log file specified, use minimal console logging
            logging.basicConfig(level=log_level)
            
    def _resolve_network_object(self, fmc_client, obj_id: str, obj_type: str, 
                                visited: Optional[Set[str]] = None) -> List[str]:
        """
        Resolve a network object/group to its constituent IP addresses/networks.
        Handles nested groups recursively with circular reference protection.
        
        Args:
            fmc_client: FMC API client instance
            obj_id: Object ID to resolve
            obj_type: Object type (Network, NetworkGroup, Host, etc.)
            visited: Set of visited IDs to prevent circular references
            
        Returns:
            List of IP addresses/networks in CIDR notation
        """
        if visited is None:
            visited = set()
            
        # Check cache first
        if obj_id in self._network_object_cache:
            logging.debug(f"Using cached network object: {obj_id}")
            return self._network_object_cache[obj_id]
            
        # Prevent circular references
        if obj_id in visited:
            logging.warning(f"Circular reference detected for network object {obj_id}")
            return []
            
        visited.add(obj_id)
        networks = []
        
        # Direct fmcapi calls - let fmcapi handle HTTP 429 and retries automatically
        # fmcapi has built-in rate limiting with 30s sleep + automatic retry
        obj_data = None
        
        if obj_type == "NetworkGroup":
            obj = fmcapi.NetworkGroups(fmc=fmc_client)
            obj.id = obj_id
            obj.get()
            if hasattr(obj, 'literals') or hasattr(obj, 'objects'):
                # Process literals (direct IP/CIDR values)
                if hasattr(obj, 'literals'):
                    for literal in obj.literals:
                        if "value" in literal:
                            networks.append(literal["value"])
                
                # Process objects (nested references)
                if hasattr(obj, 'objects'):
                    for nested_obj in obj.objects:
                        nested_id = nested_obj.get("id")
                        nested_type = nested_obj.get("type")
                        if nested_id and nested_type:
                            # Recursive call for nested objects
                            nested_networks = self._resolve_network_object(
                                fmc_client, nested_id, nested_type, visited
                            )
                            networks.extend(nested_networks)
                            
        elif obj_type == "Network":
            obj = fmcapi.Networks(fmc=fmc_client)
            obj.id = obj_id
            obj.get()
            if hasattr(obj, 'value'):
                networks.append(obj.value)
                
        elif obj_type == "Host":
            obj = fmcapi.Hosts(fmc=fmc_client)
            obj.id = obj_id
            obj.get()
            if hasattr(obj, 'value'):
                networks.append(f"{obj.value}/32")  # Single host as /32
        else:
            logging.warning(f"Unknown network object type: {obj_type}")
                            
        logging.debug(f"Resolved network object {obj_id} to {len(networks)} networks")
            
        # Cache the result
        self._network_object_cache[obj_id] = networks
        return networks
    
    def _parse_ip_range(self, ip_range: str) -> List[str]:
        """
        Parse IP range notation (e.g., 10.1.1.1-10.1.1.10) into list of individual IPs.
        
        Args:
            ip_range: IP range in format "start_ip-end_ip"
            
        Returns:
            List of IP addresses as strings, or empty list if invalid
        """
        try:
            if '-' not in ip_range:
                return [ip_range]
            
            start_ip_str, end_ip_str = ip_range.split('-', 1)
            start_ip = ipaddress.ip_address(start_ip_str.strip())
            end_ip = ipaddress.ip_address(end_ip_str.strip())
            
            # Check if range is too large (> 256 IPs) to avoid memory issues
            ip_count = int(end_ip) - int(start_ip) + 1
            if ip_count > 256:
                logging.warning(f"IP range {ip_range} contains {ip_count} addresses, checking start and end only")
                return [str(start_ip), str(end_ip)]
            
            # Generate all IPs in range
            ips = []
            current = start_ip
            while current <= end_ip:
                ips.append(str(current))
                current += 1
            return ips
            
        except Exception as e:
            logging.warning(f"Failed to parse IP range '{ip_range}': {e}")
            return []
    
    def _ip_overlaps_with_excluded_prefixes(self, ip_or_network: str) -> bool:
        """
        Check if an IP address or network overlaps with any excluded prefix.
        Handles CIDR notation, single IPs, and IP ranges (e.g., 10.1.1.1-10.1.1.10).
        
        Args:
            ip_or_network: IP address, network in CIDR notation, or IP range
            
        Returns:
            True if overlaps with any excluded prefix, False otherwise
        """
        if not self.exclude_prefixes:
            return False
        
        # Check if it's an IP range (contains '-')
        if '-' in ip_or_network:
            ips_to_check = self._parse_ip_range(ip_or_network)
            if not ips_to_check:
                return False  # Invalid range, can't determine overlap
            
            # Check each IP in the range (or just start/end for large ranges)
            for ip_str in ips_to_check:
                try:
                    ip = ipaddress.ip_address(ip_str)
                    for excluded_prefix in self.exclude_prefixes:
                        if ip in excluded_prefix:
                            logging.debug(f"IP {ip_str} from range {ip_or_network} is in excluded prefix {excluded_prefix}")
                            return True
                except ValueError:
                    continue
            return False
            
        # Handle CIDR notation or single IP
        try:
            # Parse the IP/network
            network = ipaddress.ip_network(ip_or_network, strict=False)
            
            # Check based on match mode
            for excluded_prefix in self.exclude_prefixes:
                if self.prefix_match_mode == 'overlap':
                    # Check if networks overlap in any way (default behavior)
                    if network.overlaps(excluded_prefix):
                        logging.debug(f"Network {ip_or_network} overlaps with excluded prefix {excluded_prefix}")
                        return True
                elif self.prefix_match_mode == 'subnet':
                    # Check if network is a subnet of (contained within) the excluded prefix
                    if network.subnet_of(excluded_prefix):
                        logging.debug(f"Network {ip_or_network} is subnet of excluded prefix {excluded_prefix}")
                        return True
                    
        except ValueError as e:
            logging.warning(f"Invalid IP/network '{ip_or_network}': {e}")
            return False
            
        return False
            
    def _is_rule_in_excluded_zone(self, rule_data: Dict) -> bool:
        """
        Check if rule involves any excluded zones.
        
        Args:
            rule_data: Rule data dictionary from FMC API
            
        Returns:
            True if rule should be excluded, False otherwise
        """
        if not self.exclude_zones:
            return False
            
        # Check source zones
        if "sourceZones" in rule_data and "objects" in rule_data["sourceZones"]:
            for zone_obj in rule_data["sourceZones"]["objects"]:
                if zone_obj.get("name") in self.exclude_zones:
                    return True
                    
        # Check destination zones
        if "destinationZones" in rule_data and "objects" in rule_data["destinationZones"]:
            for zone_obj in rule_data["destinationZones"]["objects"]:
                if zone_obj.get("name") in self.exclude_zones:
                    return True
                    
        return False
    
    def _is_rule_using_excluded_prefix(self, fmc_client, rule_data: Dict) -> bool:
        """
        Check if rule involves any excluded IP prefixes in source or destination networks.
        
        Args:
            fmc_client: FMC API client instance
            rule_data: Rule data dictionary from FMC API
            
        Returns:
            True if rule should be excluded, False otherwise
        """
        if not self.exclude_prefixes:
            return False
            
        rule_name = rule_data.get("name", "Unknown")
        
        # Check both source and destination networks
        for network_type in ["sourceNetworks", "destinationNetworks"]:
            if network_type not in rule_data:
                continue
                
            networks_data = rule_data[network_type]
            
            # Check literals (direct IP/CIDR values)
            if "literals" in networks_data:
                for literal in networks_data["literals"]:
                    if "value" in literal:
                        if self._ip_overlaps_with_excluded_prefixes(literal["value"]):
                            logging.info(f"Rule '{rule_name}' uses excluded prefix in {network_type} literal: {literal['value']}")
                            return True
            
            # Check objects (named network objects/groups)
            if "objects" in networks_data:
                for obj in networks_data["objects"]:
                    obj_id = obj.get("id")
                    obj_type = obj.get("type")
                    obj_name = obj.get("name", "Unknown")
                    
                    if not obj_id or not obj_type:
                        continue
                    
                    # Special handling for "any" - typically means all IPs
                    if obj_name.lower() == "any":
                        # Behavior depends on match mode
                        if self.prefix_match_mode == 'overlap':
                            # In overlap mode, "any" overlaps with everything
                            if self.exclude_prefixes:
                                logging.info(f"Rule '{rule_name}' uses 'any' in {network_type} which overlaps with excluded prefixes (overlap mode)")
                                return True
                        elif self.prefix_match_mode == 'subnet':
                            # In subnet mode, ignore "any" - we only care about specific subnets
                            logging.debug(f"Rule '{rule_name}' uses 'any' in {network_type} - ignoring in subnet mode")
                        continue
                    
                    # Resolve the network object to actual IPs
                    try:
                        resolved_networks = self._resolve_network_object(fmc_client, obj_id, obj_type)
                        
                        # Check each resolved network against excluded prefixes
                        for network in resolved_networks:
                            if self._ip_overlaps_with_excluded_prefixes(network):
                                logging.info(f"Rule '{rule_name}' uses excluded prefix in {network_type} object '{obj_name}': {network}")
                                return True
                                
                    except Exception as e:
                        logging.error(f"Error resolving network object {obj_name} ({obj_id}): {str(e)}")
                        continue
        
        return False
        
    def _should_disable_rule(self, rule_data: Dict, current_time: str, fmc_client=None) -> tuple[bool, str]:
        """
        Determine if a rule should be disabled based on criteria.
        
        Args:
            rule_data: Rule data dictionary from FMC API
            current_time: Current timestamp string
            fmc_client: FMC API client instance (optional, required for prefix exclusion)
            
        Returns:
            Tuple of (should_disable: bool, reason: str)
        """
        rule_name = rule_data.get("name", "Unknown")
        
        # Check if rule is in excluded zone
        if self._is_rule_in_excluded_zone(rule_data):
            logging.info(f"Rule '{rule_name}' skipped - involves excluded zone")
            return False, "Rule involves excluded zone"
        
        # Check if rule uses excluded IP prefixes
        if fmc_client and self._is_rule_using_excluded_prefix(fmc_client, rule_data):
            logging.info(f"Rule '{rule_name}' skipped - involves excluded IP prefix")
            return False, "Rule involves excluded IP prefix"
            
        # Check if rule is enabled and has a matching action
        rule_action = rule_data.get("action", "")
        if not (rule_data.get("enabled") and rule_action in self.rule_actions):
            return False, f"Rule is not enabled or action '{rule_action}' not in allowed actions {self.rule_actions}"
            
        # Check comment history
        if "commentHistoryList" in rule_data:
            first_comment = rule_data["commentHistoryList"][0]
            first_comment_date = first_comment.get("date", "")
            first_comment_text = first_comment.get("comment", "")
            
            # Check for previous script comments
            if "DisabledByHitCountScript" in first_comment_text:
                return True, f"Rule previously marked by script: {first_comment_text}"
                
            # Check if rule is old (created before specified year threshold)
            try:
                year = int(first_comment_date.split("-")[0])
                if year < self.year_threshold:
                    return True, f"Rule created before {self.year_threshold} (first comment: {first_comment_date})"
            except (ValueError, IndexError):
                logging.warning(f"Could not parse date for rule '{rule_name}': {first_comment_date}")
                
        else:
            # No comment history - disable
            return True, "No comment history found"
            
        return False, "Rule does not meet disable criteria"
    
    def _get_ignore_detail(self, rule_data: Dict, reason: str) -> str:
        """
        Get detailed information about why a rule was ignored.
        
        Args:
            rule_data: Rule data dictionary from FMC API
            reason: The ignore reason string
            
        Returns:
            Detailed explanation string
        """
        rule_name = rule_data.get("name", "Unknown")
        
        # Check for excluded zone
        if "excluded zone" in reason.lower():
            zones = []
            for zone_type in ["sourceZones", "destinationZones"]:
                if zone_type in rule_data and "objects" in rule_data[zone_type]:
                    for zone_obj in rule_data[zone_type]["objects"]:
                        zone_name = zone_obj.get("name")
                        if zone_name in self.exclude_zones:
                            zones.append(f"{zone_type.replace('Zones', '')}: {zone_name}")
            return "Zones: " + ", ".join(zones) if zones else "Excluded zone found"
        
        # Check for excluded IP prefix
        if "excluded IP prefix" in reason.lower() or "excluded ip prefix" in reason.lower():
            # Try to identify which networks matched
            networks_found = []
            
            # Debug: Check if network data exists
            has_source = "sourceNetworks" in rule_data
            has_dest = "destinationNetworks" in rule_data
            
            for network_type in ["sourceNetworks", "destinationNetworks"]:
                if network_type not in rule_data:
                    continue
                    
                networks_data = rule_data.get(network_type, {})
                if not networks_data:
                    continue
                
                # Check literals
                if "literals" in networks_data and networks_data["literals"]:
                    for literal in networks_data["literals"]:
                        value = literal.get("value", "")
                        if value:
                            prefix = "src" if network_type == "sourceNetworks" else "dst"
                            networks_found.append(f"{prefix}:{value}")
                
                # Check objects
                if "objects" in networks_data and networks_data["objects"]:
                    for obj in networks_data["objects"]:
                        obj_name = obj.get("name", "")
                        if obj_name:
                            prefix = "src" if network_type == "sourceNetworks" else "dst"
                            if obj_name.lower() == "any":
                                networks_found.append(f"{prefix}:ANY")
                            else:
                                networks_found.append(f"{prefix}:{obj_name}")
            
            # Build detail string
            if networks_found:
                # Limit to first 4 items for readability
                network_str = ", ".join(networks_found[:4])
                if len(networks_found) > 4:
                    network_str += f" (+{len(networks_found)-4} more)"
                return f"mode:{self.prefix_match_mode} | {network_str}"
            else:
                # Fallback if no networks found
                return f"mode:{self.prefix_match_mode} | (network details unavailable)"
        
        # Check for action/enabled issues
        if "not enabled" in reason.lower() or "action" in reason.lower():
            rule_enabled = rule_data.get("enabled", False)
            rule_action = rule_data.get("action", "UNKNOWN")
            return f"Enabled: {rule_enabled} | Action: {rule_action} | Required actions: {', '.join(self.rule_actions)}"
        
        # Check for criteria not met (age-based)
        if "does not meet disable criteria" in reason.lower():
            # Get rule creation date from comment history
            if "commentHistoryList" in rule_data and rule_data["commentHistoryList"]:
                first_comment = rule_data["commentHistoryList"][0]
                first_comment_date = first_comment.get("date", "Unknown")
                try:
                    year = int(first_comment_date.split("-")[0])
                    return f"Rule created in {year} (threshold: before {self.year_threshold})"
                except (ValueError, IndexError):
                    return f"Rule created: {first_comment_date} | Threshold: before {self.year_threshold}"
            return f"Rule does not meet age criteria (threshold: before {self.year_threshold})"
        
        # Default detail
        return reason
        
    def analyze_and_disable_rules(self) -> Dict[str, int]:
        """
        Main method to analyze hit counts and disable unused rules.
        
        Returns:
            Dictionary with operation statistics
        """
        stats = {
            "total_rules_analyzed": 0,
            "zero_hit_rules": 0,
            "rules_disabled": 0,
            "rules_skipped": 0,
            "skipped_rules": 0,  # Rules skipped due to connection issues
            "disabled_rules_details": [],  # List to store details of disabled rules
            "ignored_rules_details": []    # List to store details of ignored rules (zero hits but not disabled)
        }
        
        current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Use file_logging parameter like the simple script to let fmcapi handle its own logging
        # This ensures HTTP 429 warnings are visible
        # No try/except wrapper - let fmcapi handle all errors internally (matches simple script)
        with fmcapi.FMC(
            host=self.host,
            username=self.username,
            password=self.password,
            autodeploy=self.autodeploy,
            limit=self.page_limit,
            file_logging=self.log_file,  # Let fmcapi handle logging (matches simple script)
            debug=self.debug,
            timeout=self.timeout
        ) as fmc_client:
            
            # Log info to file
            logging.info(f"Starting hit count analysis for device '{self.device_name}' at {current_time}")
            if self.dry_run:
                logging.info("DRY RUN MODE - No changes will be made to FMC")
                
            # Only print clean connection success to console
            print(f"\nSuccessfully connected to FMC at {self.host}")
            if self.dry_run:
                print("DRY RUN MODE - No changes will be made to FMC")
                
                # Get device information
                device = fmcapi.DeviceRecords(fmc=fmc_client, name=self.device_name)
                device.get()
                
                if not hasattr(device, 'accessPolicy') or not device.accessPolicy:
                    logging.error(f"No access policy found for device '{self.device_name}'")
                    return stats
                    
                acp_id = device.accessPolicy["id"]
                logging.info(f"Found access policy ID: {acp_id}")
                
                # Get hit counts
                hit_counter = fmcapi.HitCounts(
                    fmc=fmc_client, 
                    acp_id=acp_id, 
                    device_name=self.device_name
                )
                
                hit_count_result = hit_counter.get()
                
                if "items" not in hit_count_result:
                    logging.error("No hit count data received from FMC")
                    return stats
                    
                # Identify rules with zero hits
                zero_hit_rule_ids = []
                for item in hit_count_result["items"]:
                    stats["total_rules_analyzed"] += 1
                    rule_name = item["rule"]["name"]
                    rule_type = item["rule"].get("type", "Unknown")
                    hit_count = item["hitCount"]
                    
                    logging.debug(f"Rule '{rule_name}' has {hit_count} hits")
                    
                    if hit_count == 0:
                        # Only process AccessRule types, skip default actions and other special rule types
                        if rule_type == "AccessRule":
                            zero_hit_rule_ids.append(item["rule"]["id"])
                            stats["zero_hit_rules"] += 1
                        else:
                            logging.debug(f"Skipping rule '{rule_name}' with type '{rule_type}' - not a regular access rule")
                        
                logging.info(f"Found {len(zero_hit_rule_ids)} rules with zero hit counts")
                total_to_process = min(len(zero_hit_rule_ids), self.max_rules_to_disable)
                # Only print minimal info to console - zero hit rules found and starting progress
                print(f"\nFound {len(zero_hit_rule_ids)} rules with zero hit counts")
                print(f"Processing {total_to_process} rules...")
                
                # Process each zero-hit rule
                disabled_count = 0
                processed_count = 0
                
                # Track consecutive retries across all rules
                consecutive_retries = 0
                max_consecutive_retries = 10  # Maximum number of consecutive retries allowed
                
                for rule_id in zero_hit_rule_ids:
                    if disabled_count >= self.max_rules_to_disable:
                        logging.info(f"Reached maximum rule disable limit: {self.max_rules_to_disable}")
                        break
                    
                    # Exit if too many consecutive retries
                    if consecutive_retries >= max_consecutive_retries:
                        logging.error(f"Reached maximum consecutive retries limit: {max_consecutive_retries}. Stopping processing.")
                        print(f"\nReached maximum consecutive retries limit: {max_consecutive_retries}. Stopping processing.")
                        break
                    
                    # Update progress bar
                    processed_count += 1
                    print_progress_bar(processed_count, total_to_process, prefix='Progress:', 
                                      suffix=f'({disabled_count} rules disabled)', length=50)
                    
                    # Get detailed rule information with connection retry
                    # HTTP 429 is handled by fmcapi internally
                    # But connection timeouts need manual retry with progressive backoff
                    retry_delays = [60, 90, 120, 240]  # Progressive backoff delays
                    max_retries = len(retry_delays)
                    retry_occurred = False
                    for attempt in range(max_retries):
                        try:
                            access_rule = fmcapi.AccessRules(
                                fmc=fmc_client, 
                                acp_id=acp_id, 
                                id=rule_id
                            )
                            rule_data = access_rule.get()
                            # Success - reset consecutive retries counter
                            consecutive_retries = 0
                            retry_occurred = False
                            break  # Success
                        except requests.exceptions.ConnectTimeout:
                            # Mark that a retry was needed
                            retry_occurred = True
                            consecutive_retries += 1
                            
                            if attempt < max_retries - 1:  # Don't sleep on last attempt
                                delay = retry_delays[attempt]
                                # Log details to file
                                logging.warning(f"Connection timeout for rule ID {rule_id}. Retrying in {delay}s (attempt {attempt+1}/{max_retries}, consecutive: {consecutive_retries}/{max_consecutive_retries})...")
                                
                                retry_num = attempt + 1
                                
                                # Countdown timer integrated with progress bar
                                for remaining in range(delay, 0, -1):
                                    # Show the progress bar with retry information
                                    retry_info = f"Retry #{retry_num}/{max_retries} - {remaining}s remaining (consecutive: {consecutive_retries}/{max_consecutive_retries})"
                                    print_progress_bar(
                                        processed_count, 
                                        total_to_process, 
                                        prefix='Progress:', 
                                        suffix=f'({disabled_count} rules) | {retry_info}', 
                                        length=50
                                    )
                                    time.sleep(1)
                                
                                # After countdown, restore the regular progress bar
                                print_progress_bar(processed_count, total_to_process, prefix='Progress:', 
                                                 suffix=f'({disabled_count} rules disabled)', length=50)
                            continue
                    
                    # If retries were exhausted but we want to continue with next rule
                    if retry_occurred and attempt == max_retries - 1:
                        logging.warning(f"All retries exhausted for rule ID {rule_id}. Skipping this rule and continuing with next.")
                        stats["skipped_rules"] += 1
                        # Update progress bar to show we're continuing despite retries
                        print_progress_bar(processed_count, total_to_process, prefix='Progress:', 
                                          suffix=f'({disabled_count} rules, {stats["skipped_rules"]} skipped) | Rule skipped after max retries', length=50)
                        time.sleep(2)  # Brief pause to show message
                        continue  # Skip to next rule
                            
                    # Check if API call failed (returns None on auth failure)
                    if rule_data is None:
                        logging.error(f"Failed to fetch rule {rule_id} - API returned None (possible authentication failure)")
                        logging.error("Exiting to prevent further errors.")
                        break  # Exit the loop, don't continue processing
                    
                    rule_name = rule_data.get("name", "Unknown")
                    
                    # Extract first comment if available (needed for both disabled and ignored rules)
                    first_comment = ""
                    if "commentHistoryList" in rule_data and rule_data["commentHistoryList"]:
                        first_comment_data = rule_data["commentHistoryList"][0]
                        first_comment = first_comment_data.get("comment", "")
                        first_comment_date = first_comment_data.get("date", "")
                        if first_comment and first_comment_date:
                            first_comment = f"{first_comment} ({first_comment_date})"
                    
                    should_disable, reason = self._should_disable_rule(rule_data, current_time, fmc_client)
                    
                    if should_disable:
                        # Store rule details for summary
                        rule_details = {
                            "name": rule_name,
                            "id": rule_id,
                            "first_comment": first_comment or "No comment history",
                            "reason": reason
                        }
                        
                        if self.dry_run:
                            logging.info(f"[DRY RUN] Would disable rule '{rule_name}' - {reason}")
                            stats["rules_disabled"] += 1
                            stats["disabled_rules_details"].append(rule_details)
                        else:
                            # Disable the rule and add comment
                            access_rule.enabled = False
                            comment = f"DisabledByHitCountScript {current_time} - {reason}"
                            access_rule.new_comments(action="add", value=comment)
                            access_rule.post()
                            
                            # Full details in log file, minimal console output
                            logging.info(f"Disabled rule '{rule_name}' - {reason}")
                            stats["rules_disabled"] += 1
                            stats["disabled_rules_details"].append(rule_details)
                            
                        disabled_count += 1
                    else:
                        # Log skip reason but don't clutter console
                        logging.debug(f"Skipped rule '{rule_name}' - {reason}")
                        stats["rules_skipped"] += 1
                        
                        # Store details of ignored rules (zero hits but not disabled)
                        ignored_rule_details = {
                            "name": rule_name,
                            "id": rule_id,
                            "first_comment": first_comment or "No comment history",
                            "ignore_reason": reason,
                            "ignore_detail": self._get_ignore_detail(rule_data, reason)
                        }
                        stats["ignored_rules_details"].append(ignored_rule_details)
                        
            logging.info("Hit count analysis completed")
            
            # Print final summary to console
            print("\n" + "=" * 60)
            print(f"ANALYSIS COMPLETE - Summary:")
            print(f"  - Total rules analyzed:  {stats['total_rules_analyzed']}")
            print(f"  - Rules with zero hits:  {stats['zero_hit_rules']}")
            print(f"  - Rules disabled:        {stats['rules_disabled']}")
            print(f"  - Rules skipped:         {stats['rules_skipped']}")
            print(f"  - Connection failures:   {stats['skipped_rules']}")
            if self.dry_run:
                print("\nDRY RUN COMPLETED - No changes were made to FMC.")
            print("=" * 60)
            
        return stats


def export_to_excel(stats: Dict, device_name: str, excel_file: str, dry_run: bool = False) -> None:
    """
    Export operation statistics to an Excel file with multiple sheets.
    
    Args:
        stats: Dictionary containing operation statistics
        device_name: Name of the device analyzed
        excel_file: Path to the Excel file to create
        dry_run: Whether this was a dry run
    """
    if not EXCEL_AVAILABLE:
        logging.error("openpyxl is not installed. Cannot create Excel report.")
        logging.error("Install it with: pip install openpyxl")
        return
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Create workbook
        wb = Workbook()
        
        # === Sheet 1: Operation Summary ===
        ws_summary = wb.active
        ws_summary.title = "Operation Summary"
        
        # Header
        ws_summary['A1'] = "FMC Hit Count Analysis - Operation Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        
        # Summary data
        row = 3
        ws_summary[f'A{row}'] = "Device Name:"
        ws_summary[f'B{row}'] = device_name
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "Analysis Date:"
        ws_summary[f'B{row}'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "Mode:"
        ws_summary[f'B{row}'] = "DRY RUN (No changes made)" if dry_run else "LIVE EXECUTION"
        ws_summary[f'A{row}'].font = Font(bold=True)
        if dry_run:
            ws_summary[f'B{row}'].font = Font(color="FF0000", bold=True)
        
        row += 2
        ws_summary[f'A{row}'] = "Statistics"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        ws_summary[f'A{row}'] = "Total Rules Analyzed:"
        ws_summary[f'B{row}'] = stats['total_rules_analyzed']
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "Rules with Zero Hits:"
        ws_summary[f'B{row}'] = stats['zero_hit_rules']
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "Rules Disabled:"
        ws_summary[f'B{row}'] = stats['rules_disabled']
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'].font = Font(color="00B050", bold=True)
        
        row += 1
        ws_summary[f'A{row}'] = "Rules Skipped/Ignored:"
        ws_summary[f'B{row}'] = stats['rules_skipped']
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'].font = Font(color="FF9900", bold=True)
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 50
        
        # === Sheet 2: Disabled Rules ===
        ws_disabled = wb.create_sheet("Disabled Rules")
        
        # Headers
        headers = ["Rule Name", "Rule ID", "First Comment", "Disable Reason"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_disabled.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        for row_num, rule in enumerate(stats.get('disabled_rules_details', []), 2):
            ws_disabled.cell(row=row_num, column=1, value=rule['name'])
            ws_disabled.cell(row=row_num, column=2, value=rule['id'])
            ws_disabled.cell(row=row_num, column=3, value=rule['first_comment'])
            ws_disabled.cell(row=row_num, column=4, value=rule['reason'])
            
            # Wrap text for long comments
            ws_disabled.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True)
            ws_disabled.cell(row=row_num, column=4).alignment = Alignment(wrap_text=True)
        
        # Adjust column widths
        ws_disabled.column_dimensions['A'].width = 25
        ws_disabled.column_dimensions['B'].width = 38
        ws_disabled.column_dimensions['C'].width = 50
        ws_disabled.column_dimensions['D'].width = 50
        
        # Freeze header row
        ws_disabled.freeze_panes = 'A2'
        
        # === Sheet 3: Ignored Rules ===
        ws_ignored = wb.create_sheet("Ignored Rules")
        
        # Headers
        headers = ["Rule Name", "Rule ID", "First Comment", "Ignore Reason", "Ignore Detail"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_ignored.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        for row_num, rule in enumerate(stats.get('ignored_rules_details', []), 2):
            ws_ignored.cell(row=row_num, column=1, value=rule['name'])
            ws_ignored.cell(row=row_num, column=2, value=rule['id'])
            ws_ignored.cell(row=row_num, column=3, value=rule['first_comment'])
            ws_ignored.cell(row=row_num, column=4, value=rule['ignore_reason'])
            ws_ignored.cell(row=row_num, column=5, value=rule['ignore_detail'])
            
            # Wrap text for long comments
            for col in [3, 4, 5]:
                ws_ignored.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)
        
        # Adjust column widths
        ws_ignored.column_dimensions['A'].width = 25
        ws_ignored.column_dimensions['B'].width = 38
        ws_ignored.column_dimensions['C'].width = 40
        ws_ignored.column_dimensions['D'].width = 35
        ws_ignored.column_dimensions['E'].width = 45
        
        # Freeze header row
        ws_ignored.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(excel_file)
        logging.info(f"Excel report saved to: {excel_file}")
        print(f"\nExcel report saved to: {excel_file}")
        
    except Exception as e:
        logging.error(f"Error creating Excel report: {str(e)}")
        print(f"Error creating Excel report: {str(e)}")


def format_disabled_rules_table(disabled_rules: List[Dict]) -> str:
    """
    Format disabled rules data into a readable table.
    
    Args:
        disabled_rules: List of disabled rule dictionaries
        
    Returns:
        Formatted table string
    """
    if not disabled_rules:
        return "No rules were disabled."
    
    # Calculate column widths
    name_width = max(len("Rule Name"), max(len(rule["name"]) for rule in disabled_rules))
    id_width = len("ECF40C21-3F6A-0ed3-0000-000268479583")  # Fixed width for rule IDs
    comment_width = max(len("First Comment"), max(len(rule["first_comment"]) for rule in disabled_rules))
    reason_width = max(len("Disable Reason"), max(len(rule["reason"]) for rule in disabled_rules))
    
    # Limit column widths for readability
    name_width = min(name_width, 35)
    comment_width = min(comment_width, 55)
    reason_width = min(reason_width, 45)
    
    # Create table format
    separator = "+" + "-" * (name_width + 2) + "+" + "-" * (id_width + 2) + "+" + "-" * (comment_width + 2) + "+" + "-" * (reason_width + 2) + "+"
    header_format = f"| {{:<{name_width}}} | {{:<{id_width}}} | {{:<{comment_width}}} | {{:<{reason_width}}} |"
    row_format = f"| {{:<{name_width}.{name_width}}} | {{:<{id_width}}} | {{:<{comment_width}.{comment_width}}} | {{:<{reason_width}.{reason_width}}} |"
    
    # Build table
    lines = [
        separator,
        header_format.format("Rule Name", "Rule ID", "First Comment", "Disable Reason"),
        separator
    ]
    
    for rule in disabled_rules:
        lines.append(row_format.format(
            rule["name"],
            rule["id"], 
            rule["first_comment"],
            rule["reason"]
        ))
    
    lines.append(separator)
    
    return "\n".join(lines)


def parse_arguments() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="FMC Access Rule Hit Count Analysis and Auto-Disable Script",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --host 192.168.1.100 --username admin --password secret --device firewall01.example.local
  %(prog)s --host 192.168.1.100 --username admin --password secret --device firewall01.example.local --dry-run
  %(prog)s --host 192.168.1.100 --username admin --password secret --device firewall01.example.local --max-rules 500 --exclude-zones TRUSTED DMZ
  %(prog)s --host 192.168.1.100 --username admin --password secret --device firewall01.example.local --year-threshold 2022 --rule-actions ALLOW BLOCK
  %(prog)s --host 192.168.1.100 --username admin --password secret --device firewall01.example.local --rule-actions BLOCK --dry-run
  %(prog)s --host 192.168.1.100 --username admin --password secret --device firewall01.example.local --exclude-prefixes 10.0.0.0/8 192.168.0.0/16 --dry-run
  %(prog)s --host 192.168.1.100 --username admin --password secret --device firewall01.example.local --exclude-zones TRUSTED --exclude-prefixes 172.16.0.0/12 --dry-run
  %(prog)s --host 192.168.1.100 --username admin --password secret --device firewall01.example.local --exclude-prefixes 10.2.0.0/16 --prefix-match-mode subnet --dry-run
  %(prog)s --host 192.168.1.100 --username admin --password secret --device firewall01.example.local --excel-report analysis_report.xlsx --dry-run
        """
    )
    
    # Required arguments
    required = parser.add_argument_group('required arguments')
    required.add_argument(
        '--host', 
        required=True,
        help='FMC host IP address'
    )
    required.add_argument(
        '--username',
        required=True, 
        help='FMC username'
    )
    required.add_argument(
        '--password',
        required=True,
        help='FMC password'
    )
    required.add_argument(
        '--device',
        required=True,
        dest='device_name',
        help='Target device name (cluster name)'
    )
    
    # Optional arguments
    parser.add_argument(
        '--autodeploy',
        action='store_true',
        default=False,
        help='Enable automatic deployment (default: False)'
    )
    parser.add_argument(
        '--page-limit',
        type=int,
        default=500,
        help='API page limit for queries (default: 500)'
    )
    parser.add_argument(
        '--debug',
        action='store_true',
        default=False,
        help='Enable debug logging (default: False)'
    )
    parser.add_argument(
        '--timeout',
        type=int,
        default=10,
        help='API timeout in seconds (default: 10)'
    )
    parser.add_argument(
        '--log-file',
        help='Log file name (default: log to console only)'
    )
    parser.add_argument(
        '--max-rules',
        type=int,
        default=1000,
        dest='max_rules_to_disable',
        help='Maximum number of rules to disable per run (default: 1000)'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        default=False,
        help='Simulate actions without making changes to FMC (default: False)'
    )
    parser.add_argument(
        '--exclude-zones',
        nargs='*',
        default=None,
        help='List of zones to exclude from processing (space-separated). No default - pass the flag to exclude zones.'
    )
    parser.add_argument(
        '--year-threshold',
        type=int,
        default=datetime.datetime.now().year - 1,
        help=f'Consider rules created before this year for disabling (default: {datetime.datetime.now().year - 1})'
    )
    parser.add_argument(
        '--rule-actions',
        nargs='*',
        choices=['ALLOW', 'BLOCK'],
        default=['ALLOW'],
        help='Rule actions to consider for disabling (default: ALLOW)'
    )
    parser.add_argument(
        '--exclude-prefixes',
        nargs='*',
        default=None,
        help='IP prefixes to exclude from processing (CIDR notation, space-separated, e.g., 10.0.0.0/8 192.168.1.0/24). Rules using these IPs/networks will be skipped.'
    )
    parser.add_argument(
        '--prefix-match-mode',
        choices=['overlap', 'subnet'],
        default='overlap',
        help='Mode for matching excluded prefixes: "overlap" (default) excludes rules with any network overlap (including supersets like "any"), "subnet" only excludes rules with networks that are subsets of excluded prefixes'
    )
    parser.add_argument(
        '--excel-report',
        help='Generate an Excel report file with operation summary, disabled rules, and ignored rules (e.g., report.xlsx). Requires openpyxl package.'
    )
    
    return parser.parse_args()


def main() -> int:
    """Main function."""
    args = parse_arguments()
    
    # Create FMC Rule Manager instance
    manager = FMCRuleManager(
        host=args.host,
        username=args.username,
        password=args.password,
        device_name=args.device_name,
        autodeploy=args.autodeploy,
        page_limit=args.page_limit,
        debug=args.debug,
        timeout=args.timeout,
        log_file=args.log_file,
        max_rules_to_disable=args.max_rules_to_disable,
        dry_run=args.dry_run,
        exclude_zones=args.exclude_zones,
        year_threshold=args.year_threshold,
        rule_actions=args.rule_actions,
        exclude_prefixes=args.exclude_prefixes,
        prefix_match_mode=args.prefix_match_mode
    )
    
    # Run the analysis - let any exceptions propagate (matches simple script)
    stats = manager.analyze_and_disable_rules()
    
    # Print summary
    print("\n" + "="*50)
    print("OPERATION SUMMARY")
    print("="*50)
    print(f"Total rules analyzed: {stats['total_rules_analyzed']}")
    print(f"Rules with zero hits: {stats['zero_hit_rules']}")
    print(f"Rules disabled: {stats['rules_disabled']}")
    print(f"Rules skipped: {stats['rules_skipped']}")
    if args.dry_run:
        print("\nNOTE: This was a dry run - no changes were made to FMC")
    
    # Disabled rules table is now only in Excel report, not in log or console
    print("="*50)
    
    # Generate Excel report if requested
    if args.excel_report:
        export_to_excel(stats, args.device_name, args.excel_report, args.dry_run)
    
    return 0


if __name__ == "__main__":
    sys.exit(main())